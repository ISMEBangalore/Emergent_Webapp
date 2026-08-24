"""
Weekly CRM report engine (vectorised).

Parses a CRM Lead dump (.xlsx) and optional Application dumps, and computes the
Program x Lead-Stage matrix plus derived metrics driven by configurable rules.
"""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd


DEFAULT_SETTINGS: Dict[str, Any] = {
    "programs": ["B.Com", "BBA", "PGDM"],
    "program_aliases": {},              # e.g. {"PGDM": ["PGDM(MKT/FIN/HR/BA/IA)"]}
    "stage_rows": [
        "APPLIED",
        "COLD Unverified leads",
        "COLD Verified leads (includes NI)",
        "JOINED IN ANOTHER COLLEGE",
        "UNANSWERED 3 TIMES",
        "JUNK",
        "NOT ELIGIBLE",
        "NOT REACHABLE/SWITCH OFF",
        "Untouched",
        "WARM",
    ],
    "verified_logic": "any",           # any | all | mobile | email
    "relevant_stages": ["WARM", "HOT", "APPLIED", "INTERESTED IN PGDM R&L"],
    "api_patterns": ["API"],
    "redirect_patterns": ["REDIRECT", "PUSH", "WIDGET"],
    "application_code_field": "Agent Code",
    "exclude_test_leads": True,
    "test_keywords": ["test"],
    "excluded_publishers": [],
    "included_publishers": [],
    "applications_payment_approved_only": True,
}

PROGRAM_CANDIDATES = ["Course", "Programme", "Program", "Form Interested In"]
STAGE_CANDIDATES = ["Lead Stage", "First Lead Stage"]
EMAIL_VERIF_CANDIDATES = ["Email Verification Status"]
MOBILE_VERIF_CANDIDATES = ["Mobile Verification Status"]
ORIGIN_CANDIDATES = [
    "Lead Origin(Primary)", "Primary Traffic Channel", "Lead Origin",
    "Latest Traffic Channel", "Lead Type(Secondary)",
]
WIDGET_CANDIDATES = ["Widget Name"]
PAYMENT_CANDIDATES = ["Payment Approved"]
PUBLISHER_CANDIDATES = ["Publisher Name", "Publisher", "Publisher(Primary)"]
STATE_CANDIDATES = ["State", "Registered State"]
CITY_CANDIDATES = ["City", "Registered City"]
# The CRM's City field occasionally holds a district name rather than an actual city
# (mainly for Delhi, which has no single "city" — it's NCT split into districts — and
# a stray "Bengaluru Urban" district value). Folded into the city they actually mean.
CITY_ALIASES = {
    "BENGALURU URBAN": "BENGALURU", "BENGALURU RURAL": "BENGALURU",
    "NEW DELHI": "DELHI", "NORTH DELHI": "DELHI", "SOUTH DELHI": "DELHI",
    "EAST DELHI": "DELHI", "WEST DELHI": "DELHI", "CENTRAL DELHI": "DELHI",
    "NORTH WEST DELHI": "DELHI", "NORTH EAST DELHI": "DELHI",
    "SOUTH WEST DELHI": "DELHI", "SOUTH EAST DELHI": "DELHI", "SHAHDARA": "DELHI",
}
DATE_CANDIDATES = [
    "User Registration Date", "Registration Date", "Created On", "Lead Created Date",
    "Lead Creation Date", "Created Date", "Creation Date", "Lead Created On", "Created",
    "Enquiry Date", "Date Created", "Lead Date",
]


def read_data_sheet(content: bytes) -> pd.DataFrame:
    """Read the sheet that actually holds row-level data.
    Some CRM exports put a pivot/summary sheet first; pick the largest sheet."""
    import openpyxl
    best = 0
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        best_score = -1
        for s in wb.sheetnames:
            ws = wb[s]
            score = (ws.max_row or 0) * (ws.max_column or 0)
            if score > best_score:
                best, best_score = s, score
        wb.close()
    except Exception:
        best = 0
    # calamine (Rust) parses large .xlsx files roughly 5-6x faster than openpyxl's
    # pure-Python parser — meaningful when uploads run 90-100MB.
    return pd.read_excel(io.BytesIO(content), engine="calamine", sheet_name=best)


def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _find_col(df: pd.DataFrame, candidates: List[str], prefer_data: bool = False) -> Optional[str]:
    lut = {_norm(c): c for c in df.columns}
    matches: List[str] = []
    for cand in candidates:
        key = _norm(cand)
        if key in lut and lut[key] not in matches:
            matches.append(lut[key])
    for cand in candidates:
        key = _norm(cand)
        for k, orig in lut.items():
            if key and key in k and orig not in matches:
                matches.append(orig)
    if not matches:
        return None
    if prefer_data:
        for m in matches:
            if df[m].notna().any():
                return m
    return matches[0]


def _alnum(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def program_merge_targets(programs: List[str]) -> Dict[str, str]:
    """Maps each configured program to the shortest sibling program its raw name
    is a variant/specialization spelling of - e.g. "PGDM (MKT/FIN/HR/BA/IA)" ->
    "PGDM" - so a course text like that automatically folds into the base
    program everywhere (main report, Application Insight, Geography, Verified
    Lead Analysis) without needing it hand-configured as a program_alias. A
    program with no such sibling maps to itself."""
    keys = {p: _alnum(p) for p in programs}
    target: Dict[str, str] = {}
    for p in programs:
        siblings = [q for q in programs if q != p and keys[q] and keys[p].startswith(keys[q])]
        target[p] = min(siblings, key=lambda q: len(keys[q])) if siblings else p
    return target


def program_series(series: pd.Series, programs: List[str],
                   aliases: Dict[str, List[str]] = None) -> pd.Series:
    """Map a Course/Programme column to one of the configured program names.
    Exact (alnum) match first; substring fallback only for match keys that have
    no exact match anywhere (e.g. short canonical names like B.Com/PGDM),
    so selecting a full raw course name stays precise and never double-counts.

    `aliases` lets a program absorb other raw CRM text values that should
    count toward it (e.g. {"PGDM": ["PGDM(MKT/FIN/HR/BA/IA)"]}) without
    listing them as separate program columns."""
    aliases = aliases or {}
    up = series.astype("string").fillna("")
    key = up.str.upper().str.replace(r"[^A-Z0-9]", "", regex=True)
    out = pd.Series([None] * len(series), index=series.index, dtype="object")

    prog_match_keys: Dict[str, List[str]] = {}
    for p in programs:
        keys = [_alnum(p)] + [_alnum(a) for a in aliases.get(p, [])]
        prog_match_keys[p] = list(dict.fromkeys(k for k in keys if k))  # dedup, keep order

    for p in programs:
        for mk in prog_match_keys[p]:
            out = out.mask(out.isna() & (key == mk), p)
    for p in programs:
        for mk in prog_match_keys[p]:
            if not (key == mk).any():
                out = out.mask(out.isna() & (key != "") & key.str.contains(re.escape(mk), regex=True), p)
    return out


_GEO_UNKNOWN_TOKENS = {"", "NAN", "NONE", "N/A", "NA", "-", "NULL"}


def geo_series(series: pd.Series) -> pd.Series:
    """Normalize a State/City column: uppercase, trim, and fold the CRM's own
    'STATE NOT AVAILABLE' / 'CITY NOT AVAILABLE' placeholders (and blanks) into
    a single 'Unknown' bucket instead of scattering them as junk categories."""
    v = series.astype("string").str.strip().str.upper().fillna("")
    unknown = v.isin(_GEO_UNKNOWN_TOKENS) | v.str.contains("NOT AVAILABLE", na=False)
    return v.mask(unknown, "UNKNOWN")


def city_series(series: pd.Series) -> pd.Series:
    v = geo_series(series)
    return v.replace(CITY_ALIASES)


def _prog_from_series(s: pd.Series) -> pd.Series:
    u = s.astype("string").str.upper().fillna("")
    out = pd.Series(np.where(u.str.contains("PGDM"), "PGDM",
             np.where(u.str.contains("BBA"), "BBA",
             np.where(u.str.contains("B.COM") | u.str.contains("BCOM") | u.str.contains("B COM") | u.str.contains("B. COM"), "B.Com",
             np.where(u.str.contains("BCA"), "BCA",
             np.where(u.str.strip().isin(["", "NAN", "NONE"]), None, "Other"))))), index=s.index)
    return out


def compute_report(
    lead_bytes: bytes,
    settings: Dict[str, Any],
    amount_spent: Optional[Dict[str, float]] = None,
    additional_attributed: Optional[Dict[str, float]] = None,
    application_counts: Optional[Dict[str, Any]] = None,
    date_range: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    cfg = {**DEFAULT_SETTINGS, **(settings or {})}
    raw_programs: List[str] = cfg["programs"]
    program_aliases: Dict[str, List[str]] = cfg.get("program_aliases") or {}
    merge_target = program_merge_targets(raw_programs)
    programs: List[str] = [p for p in raw_programs if merge_target[p] == p]
    stage_rows: List[str] = cfg["stage_rows"]
    amount_spent = amount_spent or {}
    additional_attributed = additional_attributed or {}
    application_counts = application_counts or {}
    apps_by_program = application_counts.get("by_program", {})
    apps_by_publisher = {str(k).strip().upper(): v for k, v in (application_counts.get("by_publisher") or {}).items()}
    apps_by_program_publisher = {
        p: {str(k).strip().upper(): v for k, v in (pubs or {}).items()}
        for p, pubs in (application_counts.get("by_program_publisher") or {}).items()
    }
    apps_by_state = application_counts.get("by_state") or {}
    apps_by_city = application_counts.get("by_city") or {}
    apps_by_program_state = application_counts.get("by_program_state") or {}
    apps_by_program_city = application_counts.get("by_program_city") or {}
    apps_by_publisher_state = {
        str(k).strip().upper(): v for k, v in (application_counts.get("by_publisher_state") or {}).items()
    }
    apps_by_publisher_city = {
        str(k).strip().upper(): v for k, v in (application_counts.get("by_publisher_city") or {}).items()
    }
    apps_by_program_publisher_state = {
        p: {str(k).strip().upper(): v for k, v in (pubs or {}).items()}
        for p, pubs in (application_counts.get("by_program_publisher_state") or {}).items()
    }
    apps_by_program_publisher_city = {
        p: {str(k).strip().upper(): v for k, v in (pubs or {}).items()}
        for p, pubs in (application_counts.get("by_program_publisher_city") or {}).items()
    }
    _blank_app_counts = {"with_code": 0, "without_code": 0, "via_redirect": 0, "via_api": 0}

    df = read_data_sheet(lead_bytes)
    raw_n = len(df)

    # ---- Exclude TEST leads (by name / remark / email / stage) ----
    test_excluded = 0
    if cfg.get("exclude_test_leads", True):
        kws = [str(k).lower().strip() for k in cfg.get("test_keywords", ["test"]) if str(k).strip()]
        is_test = pd.Series(False, index=df.index)
        name_col = _find_col(df, ["Registered Name", "Name", "First Name"])
        rem_col = _find_col(df, ["Lead Remark", "Remark", "Remarks"])
        email_col = _find_col(df, ["Email", "Email Id", "Email Address"])
        stage_col_tmp = _find_col(df, STAGE_CANDIDATES, prefer_data=True)
        for kw in kws:
            word_pat = r"\b" + re.escape(kw) + r"\b"
            if name_col:
                is_test |= df[name_col].astype("string").str.contains(word_pat, case=False, na=False, regex=True)
            if rem_col:
                is_test |= df[rem_col].astype("string").str.contains(word_pat, case=False, na=False, regex=True)
            if email_col:
                is_test |= df[email_col].astype("string").str.contains(kw, case=False, na=False, regex=False)
        if stage_col_tmp:
            is_test |= df[stage_col_tmp].astype("string").str.strip().str.upper().eq("TEST LEADS").fillna(False)
        test_excluded = int(is_test.sum())
        if test_excluded:
            df = df[~is_test].reset_index(drop=True)

    # ---- Lead creation date coverage + optional range filter ----
    date_range = date_range or {}
    d_start = (str(date_range.get("start") or "")).strip() or None
    d_end = (str(date_range.get("end") or "")).strip() or None
    col_date = _find_col(df, DATE_CANDIDATES, prefer_data=True)
    date_min = date_max = None
    rows_before_date = len(df)
    date_filtered = 0
    if col_date is not None:
        parsed = pd.to_datetime(df[col_date], errors="coerce", dayfirst=True)
        if parsed.notna().any():
            date_min = parsed.min().strftime("%Y-%m-%d")
            date_max = parsed.max().strftime("%Y-%m-%d")
        if d_start or d_end:
            keep = pd.Series(True, index=df.index)
            if d_start:
                keep &= parsed >= pd.Timestamp(d_start)
            if d_end:
                keep &= parsed < (pd.Timestamp(d_end) + pd.Timedelta(days=1))
            keep = keep.fillna(False)
            date_filtered = int((~keep).sum())
            df = df[keep].reset_index(drop=True)
    n = len(df)

    col_prog = _find_col(df, PROGRAM_CANDIDATES, prefer_data=True)
    col_stage = _find_col(df, STAGE_CANDIDATES, prefer_data=True)
    col_email = _find_col(df, EMAIL_VERIF_CANDIDATES)
    col_mobile = _find_col(df, MOBILE_VERIF_CANDIDATES)
    col_origin = _find_col(df, ORIGIN_CANDIDATES, prefer_data=True)
    col_widget = _find_col(df, WIDGET_CANDIDATES)
    col_payment = _find_col(df, PAYMENT_CANDIDATES)
    col_agent = _find_col(df, [cfg.get("application_code_field", "Agent Code")])
    col_pub = _find_col(df, PUBLISHER_CANDIDATES, prefer_data=True)
    col_state = _find_col(df, STATE_CANDIDATES, prefer_data=True)
    col_city = _find_col(df, CITY_CANDIDATES, prefer_data=True)

    all_cols = list(programs) + ["Other"]

    # ---- Program (vectorised) ----
    if col_prog is not None:
        # Classify against every raw program spelling (including variants like
        # "PGDM (MKT/FIN/HR/BA/IA)") so those rows still match something, then fold
        # the result through merge_target so they land on the base program's column.
        prog = program_series(df[col_prog], raw_programs, program_aliases).map(lambda p: merge_target.get(p, p))
    else:
        prog = pd.Series([None] * n, index=df.index, dtype="object")
    # Fallback via widget / payment where program still unknown
    need = prog.isna()
    if need.any() and (col_widget or col_payment):
        wp = pd.Series([""] * n, index=df.index, dtype="object")
        if col_widget:
            wp = wp.str.cat(df[col_widget].astype("string").fillna("").str.upper(), sep=" ")
        if col_payment:
            wp = wp.str.cat(df[col_payment].astype("string").fillna("").str.upper(), sep=" ")
        fb = pd.Series(np.where(wp.str.contains("PGDM", regex=False) | wp.str.contains("WIDGET(PG)", regex=False) | wp.str.contains("PG-EX", regex=False), "PGDM",
                        np.where(wp.str.contains("UG", regex=False), "UG", None)), index=df.index)
        prog = prog.where(~need, fb)
    unclassified = int(prog.isna().sum())
    prog = prog.where(prog.isin(all_cols), "Other")

    # ---- Verified (vectorised) ----
    e = (df[col_email].astype("string").str.upper() == "VERIFIED") if col_email else pd.Series(False, index=df.index)
    m = (df[col_mobile].astype("string").str.upper() == "VERIFIED") if col_mobile else pd.Series(False, index=df.index)
    e = e.fillna(False); m = m.fillna(False)
    logic = cfg.get("verified_logic", "any")
    verified = {"email": e, "mobile": m, "all": e & m}.get(logic, e | m)

    # ---- Stage bucket (vectorised) ----
    if col_stage is not None:
        raw = df[col_stage].astype("string").str.strip().str.upper()
    else:
        raw = pd.Series([pd.NA] * n, index=df.index, dtype="string")
    empty = raw.isna() | (raw == "")
    bucket = pd.Series("Untouched", index=df.index, dtype="object")
    def setb(mask, label):
        bucket.loc[mask & (bucket == "Untouched")] = label
    setb(raw == "APPLIED", "APPLIED")
    setb((raw == "COLD") & verified, "COLD Verified leads (includes NI)")
    setb((raw == "COLD") & ~verified, "COLD Unverified leads")
    setb(raw.isin(["NOT INTERESTED", "FORM STARTED - NOT INTERESTED"]), "COLD Unverified leads")
    setb(raw.isin(["JOINED IN OTHER COLLEGE", "JOINED IN ANOTHER COLLEGE"]), "JOINED IN ANOTHER COLLEGE")
    setb(raw == "UNANSWERED 3 TIMES", "UNANSWERED 3 TIMES")
    setb(raw.isin(["JUNK", "WRONG NUMBER", "TEST LEADS", "DONT PURGE"]), "JUNK")
    setb(raw.isin(["NOT ELIGIBLE", "ELIGIBLE FOR NEXT YEAR"]), "NOT ELIGIBLE")
    setb(raw == "NOT REACHABLE/SWITCH OFF", "NOT REACHABLE/SWITCH OFF")
    setb(raw.isin(["WARM", "HOT"]), "WARM")
    # unknown non-empty stages -> Untouched (already default); keep empty as Untouched too
    bucket.loc[empty] = "Untouched"

    if col_pub is not None:
        pub_raw = df[col_pub].astype("string").str.strip()
        pub_raw = pub_raw.where(pub_raw.notna() & (pub_raw != ""), "Unknown")
    else:
        pub_raw = pd.Series(["Unknown"] * n, index=df.index, dtype="object")

    work = pd.DataFrame({"prog": prog.values, "pub": pub_raw.values,
                         "bucket": bucket.values, "verified": verified.values})
    relevant_set = {s.upper() for s in cfg["relevant_stages"]}
    work["relevant"] = raw.isin(relevant_set).values
    work["state"] = geo_series(df[col_state]).values if col_state is not None else "UNKNOWN"
    work["city"] = city_series(df[col_city]).values if col_city is not None else "UNKNOWN"

    # ---- Channel kind (vectorised) ----
    if col_origin is not None:
        o = df[col_origin].astype("string").str.upper().fillna("")
        is_api = pd.Series(False, index=df.index)
        for p in cfg.get("api_patterns", []):
            is_api = is_api | o.str.contains(re.escape(p.upper()))
        is_redir = pd.Series(False, index=df.index)
        for p in cfg.get("redirect_patterns", []):
            is_redir = is_redir | o.str.contains(re.escape(p.upper()))
        is_redir = is_redir & ~is_api
    else:
        is_api = pd.Series(False, index=df.index)
        is_redir = pd.Series(False, index=df.index)
    work["is_api"] = is_api.values
    work["is_redir"] = is_redir.values

    def agg(frame, colname, cats):
        def counts(mask=None):
            sub = frame if mask is None else frame[mask.values if hasattr(mask, "values") else mask]
            vc = sub[colname].value_counts()
            return {c: int(vc.get(c, 0)) for c in cats}
        ct = pd.crosstab(frame["bucket"], frame[colname])
        mc = {row: {c: int(ct.loc[row, c]) if (row in ct.index and c in ct.columns) else 0 for c in cats}
              for row in stage_rows}
        return (counts(), counts(frame["verified"]), counts(frame["is_redir"]),
                counts(frame["is_redir"] & frame["verified"]), counts(frame["is_api"]),
                counts(frame["is_api"] & frame["verified"]), counts(frame["relevant"]), mc)

    (total_leads, verified_leads, redirect_leads, redirect_verified,
     api_leads, api_verified, relevant_leads, matrix_counts) = agg(work, "prog", all_cols)

    # ---- Publisher breakdown (all detected publishers, honouring include/exclude) ----
    pub_vc = work["pub"].value_counts()
    available_publishers = [{"name": str(k), "count": int(v)} for k, v in pub_vc.items()][:200]
    excluded = {str(e).strip().upper() for e in cfg.get("excluded_publishers", []) if str(e).strip()}
    included = [str(i).strip() for i in cfg.get("included_publishers", []) if str(i).strip()]
    ordered = [str(x) for x in pub_vc.index.tolist()]
    if included:
        incl_up = {i.upper() for i in included}
        ordered = [p for p in ordered if p.upper() in incl_up]
    ordered = [p for p in ordered if p.upper() not in excluded]
    pub_cats = ordered[:40] or ["Unknown"]
    if col_prog is not None:
        cvc = df[col_prog].astype("string").str.strip()
        cvc = cvc[cvc.notna() & (cvc != "")].value_counts()
        available_courses = [{"name": str(k), "count": int(v)} for k, v in cvc.items()][:80]
    else:
        available_courses = []

    def make_publisher_result(frame):
        (t, ver, rd, rdv, ap, apv, rel, mc) = agg(frame, "pub", pub_cats)
        app_counts = {c: apps_by_publisher.get(c.strip().upper(), _blank_app_counts) for c in pub_cats}
        return build_result(
            programs=pub_cats, stage_rows=stage_rows, matrix_counts=mc,
            total_leads=t, verified_leads=ver, redirect_leads=rd, redirect_verified=rdv,
            api_leads=ap, api_verified=apv, relevant_leads=rel,
            application_counts=app_counts, amount_spent={}, additional_attributed={},
            detected_columns={"publisher": col_pub},
            data_quality={"total_rows": len(frame), "publisher_column_present": bool(col_pub),
                          "publisher_count": int(len(pub_vc))},
        )

    publisher_result = make_publisher_result(work)
    # Per-program publisher breakdowns (same publisher columns for consistency)
    publisher_reports = {"All": publisher_result}
    for prog_name in programs:
        publisher_reports[prog_name] = make_publisher_result(work[work["prog"] == prog_name])

    result = build_result(
        programs=programs, stage_rows=stage_rows, matrix_counts=matrix_counts,
        total_leads=total_leads, verified_leads=verified_leads,
        redirect_leads=redirect_leads, redirect_verified=redirect_verified,
        api_leads=api_leads, api_verified=api_verified, relevant_leads=relevant_leads,
        application_counts=apps_by_program, amount_spent=amount_spent,
        additional_attributed=additional_attributed,
        detected_columns={
            "program": col_prog, "stage": col_stage,
            "email_verification": col_email, "mobile_verification": col_mobile,
            "lead_origin": col_origin, "agent_code": col_agent, "publisher": col_pub,
        },
        data_quality={
            "total_rows": n, "unclassified_program": unclassified,
            "program_column_present": bool(col_prog), "stage_column_present": bool(col_stage),
            "test_leads_excluded": test_excluded, "raw_rows": raw_n,
            "date_column": col_date, "data_date_min": date_min, "data_date_max": date_max,
            "date_filtered_out": date_filtered, "rows_before_date_filter": rows_before_date,
            "date_range": {"start": d_start or None, "end": d_end or None},
        },
    )
    result["publisher_report"] = publisher_result
    result["publisher_reports"] = publisher_reports

    # ---- Geography (State / City) ----
    def make_geo_breakdown(frame: pd.DataFrame, colname: str, app_counts: Dict[str, int]) -> Dict[str, Any]:
        g = frame.groupby(colname, observed=True)
        leads = g.size()
        verified_ct = g["verified"].sum()
        relevant_ct = g["relevant"].sum()
        out: Dict[str, Any] = {}
        for name in leads.index:
            key = str(name)
            apps = int(app_counts.get(key, 0))
            lead_n = int(leads[name])
            out[key] = {
                "leads": lead_n,
                "verified_leads": int(verified_ct[name]),
                "relevant_leads": int(relevant_ct[name]),
                "applications": apps,
                "conversion_pct": round(apps / lead_n * 100, 2) if lead_n else None,
            }
        # Locations with applications but zero matching leads (e.g. a state present
        # only in the application file) still deserve a row.
        for key, apps in app_counts.items():
            if key not in out and apps:
                out[key] = {"leads": 0, "verified_leads": 0, "relevant_leads": 0,
                           "applications": int(apps), "conversion_pct": None}
        return out

    # Nested [program_or_All][publisher_or_All] -> {location: {...}} so the Geo tab
    # can slice by Program, by Publisher, or by both at once (a publisher's reach
    # within one program) without three separate parallel structures.
    geo_by_state: Dict[str, Dict[str, Any]] = {}
    geo_by_city: Dict[str, Dict[str, Any]] = {}

    def add_geo(prog_key: str, pub_key: str, frame: pd.DataFrame,
               state_counts: Dict[str, int], city_counts: Dict[str, int]) -> None:
        geo_by_state.setdefault(prog_key, {})[pub_key] = make_geo_breakdown(frame, "state", state_counts)
        geo_by_city.setdefault(prog_key, {})[pub_key] = make_geo_breakdown(frame, "city", city_counts)

    add_geo("All", "All", work, apps_by_state, apps_by_city)
    for prog_name in programs:
        sub = work[work["prog"] == prog_name]
        add_geo(prog_name, "All", sub, apps_by_program_state.get(prog_name, {}), apps_by_program_city.get(prog_name, {}))
    for pub_name in pub_cats:
        pub_key = pub_name.strip().upper()
        pub_frame = work[work["pub"] == pub_name]
        add_geo("All", pub_name, pub_frame, apps_by_publisher_state.get(pub_key, {}), apps_by_publisher_city.get(pub_key, {}))
        for prog_name in programs:
            sub = pub_frame[pub_frame["prog"] == prog_name]
            add_geo(prog_name, pub_name, sub,
                    apps_by_program_publisher_state.get(prog_name, {}).get(pub_key, {}),
                    apps_by_program_publisher_city.get(prog_name, {}).get(pub_key, {}))
    result["geo_by_state"] = geo_by_state
    result["geo_by_city"] = geo_by_city
    result["data_quality"]["state_column_present"] = bool(col_state)
    result["data_quality"]["city_column_present"] = bool(col_city)
    result["data_quality"]["available_courses"] = available_courses
    result["data_quality"]["available_publishers"] = available_publishers

    # Program breakdown per publisher (columns = programs, one report per publisher)
    def make_program_result(frame, pub_name):
        (t, ver, rd, rdv, ap, apv, rel, mc) = agg(frame, "prog", all_cols)
        pub_key = pub_name.strip().upper()
        app_counts = {p: apps_by_program_publisher.get(p, {}).get(pub_key, _blank_app_counts) for p in programs}
        return build_result(
            programs=programs, stage_rows=stage_rows, matrix_counts=mc,
            total_leads=t, verified_leads=ver, redirect_leads=rd, redirect_verified=rdv,
            api_leads=ap, api_verified=apv, relevant_leads=rel,
            application_counts=app_counts, amount_spent={}, additional_attributed={},
            detected_columns={"publisher": col_pub}, data_quality={"total_rows": len(frame)},
        )

    program_reports = {"All": {k: result[k] for k in
                        ("programs", "columns", "matrix", "summary", "detected_columns", "data_quality")}}
    for pub_name in pub_cats:
        program_reports[pub_name] = make_program_result(work[work["pub"] == pub_name], pub_name)
    result["program_reports"] = program_reports
    return result


def build_result(programs, stage_rows, matrix_counts, total_leads, verified_leads,
                 redirect_leads, redirect_verified, api_leads, api_verified, relevant_leads,
                 application_counts, amount_spent, additional_attributed,
                 detected_columns, data_quality):
    """Build the report result dict from pre-aggregated count dictionaries."""
    def totrow(d):
        return sum(d.get(p, 0) for p in programs)

    apps_with = {p: int(application_counts.get(p, {}).get("with_code", 0)) for p in programs}
    apps_without = {p: int(application_counts.get(p, {}).get("without_code", 0)) for p in programs}
    apps_redirect = {p: int(application_counts.get(p, {}).get("via_redirect", 0)) for p in programs}
    apps_api = {p: int(application_counts.get(p, {}).get("via_api", 0)) for p in programs}
    total_apps = {p: apps_with[p] + apps_without[p] for p in programs}

    def pct(num, den):
        return round(num / den * 100, 2) if den else None

    grand_total = totrow(total_leads)
    result: Dict[str, Any] = {
        "programs": programs,
        "columns": programs + ["Total"],
        "detected_columns": detected_columns,
        "data_quality": data_quality,
        "matrix": [],
        "summary": [],
    }

    for row in stage_rows:
        entry = {"stage": row, "values": {}, "pct": {}, "total": totrow(matrix_counts[row])}
        for p in programs:
            entry["values"][p] = matrix_counts[row][p]
            entry["pct"][p] = pct(matrix_counts[row][p], total_leads[p])
        entry["total_pct"] = pct(entry["total"], grand_total)
        result["matrix"].append(entry)

    def srow(label, values, pcts=None, fmt="int", kind="metric"):
        e = {"label": label, "fmt": fmt, "kind": kind, "values": {}, "pct": {}}
        tot = 0
        for p in programs:
            v = values.get(p, 0)
            e["values"][p] = v
            tot += v or 0
            if pcts is not None:
                e["pct"][p] = pcts.get(p)
        e["total"] = round(tot, 2)
        if pcts is not None:
            e["total_pct"] = pcts.get("Total")
        return e

    S = result["summary"]
    S.append(srow("Total Leads", total_leads, {**{p: 100.0 for p in programs}, "Total": 100.0}, kind="header"))
    S.append(srow("Verified Leads", verified_leads))
    S.append(srow("Verified leads (%) of Total Leads", verified_leads,
                  {**{p: pct(verified_leads[p], total_leads[p]) for p in programs}, "Total": pct(totrow(verified_leads), grand_total)}, fmt="pct_only"))
    S.append(srow("Total Redirect Leads", redirect_leads))
    S.append(srow("Total Verified redirect leads", redirect_verified))
    S.append(srow("Verified % of Redirect Leads", redirect_verified,
                  {**{p: pct(redirect_verified[p], redirect_leads[p]) for p in programs}, "Total": pct(totrow(redirect_verified), totrow(redirect_leads))}, fmt="pct_only"))
    S.append(srow("Total API Leads", api_leads))
    S.append(srow("Total Verified API leads", api_verified))
    S.append(srow("Verified % of API Leads", api_verified,
                  {**{p: pct(api_verified[p], api_leads[p]) for p in programs}, "Total": pct(totrow(api_verified), totrow(api_leads))}, fmt="pct_only"))
    S.append(srow("No. of Applications with codes", apps_with))
    S.append(srow("No. of Applications without codes", apps_without))
    S.append({"label": "Lead Analysis", "kind": "section"})
    S.append(srow("Relevant Leads", relevant_leads))
    S.append(srow("% Relevant Leads", relevant_leads,
                  {**{p: pct(relevant_leads[p], total_leads[p]) for p in programs}, "Total": pct(totrow(relevant_leads), grand_total)}, fmt="pct_only"))
    S.append(srow("Total No. of Applications", total_apps))
    S.append(srow("% of Applications", total_apps,
                  {**{p: pct(total_apps[p], total_leads[p]) for p in programs}, "Total": pct(sum(total_apps.values()), grand_total)}, fmt="pct_only"))
    S.append(srow("No of applications through redirect leads", apps_redirect))
    S.append(srow("% of application based on redirect leads", apps_redirect,
                  {**{p: pct(apps_redirect[p], total_apps[p]) for p in programs}, "Total": pct(sum(apps_redirect.values()), sum(total_apps.values()))}, fmt="pct_only"))
    S.append(srow("No of applications through API leads", apps_api))
    S.append(srow("% of application based on API leads", apps_api,
                  {**{p: pct(apps_api[p], total_apps[p]) for p in programs}, "Total": pct(sum(apps_api.values()), sum(total_apps.values()))}, fmt="pct_only"))

    spent = {p: float(amount_spent.get(p, 0) or 0) for p in programs}
    add_attr = {p: float(additional_attributed.get(p, 0) or 0) for p in programs}
    S.append(srow("Amount Spent", spent, fmt="money"))
    cost_per_app = {p: round(spent[p] / total_apps[p], 2) if total_apps[p] else 0 for p in programs}
    S.append(srow("Cost/Application", cost_per_app, fmt="money"))
    S.append(srow("Additional Attributed Applications", add_attr, fmt="int"))
    mod_cost = {p: round(spent[p] / (total_apps[p] + add_attr[p]), 2) if (total_apps[p] + add_attr[p]) else 0 for p in programs}
    S.append(srow("Modified CPA after attribution", mod_cost, fmt="money"))

    return result



def _extract_raw_counts(result: Dict[str, Any], columns: List[str], stage_rows: List[str]) -> Dict[str, Any]:
    """Recovers the raw count dicts build_result() was originally built from, by
    reading them back out of its summary/matrix rows. Lets a built result be
    re-diffed and rebuilt (via _diff_result) without duplicating build_result()'s
    percentage/total logic."""
    summ = {s["label"]: s.get("values", {}) for s in result.get("summary", []) if "values" in s}
    matrix = {m["stage"]: m.get("values", {}) for m in result.get("matrix", [])}

    def vals(label):
        return dict(summ.get(label, {}))

    with_code, without_code = vals("No. of Applications with codes"), vals("No. of Applications without codes")
    via_redirect, via_api = vals("No of applications through redirect leads"), vals("No of applications through API leads")
    return {
        "matrix_counts": {row: dict(matrix.get(row, {})) for row in stage_rows},
        "total_leads": vals("Total Leads"), "verified_leads": vals("Verified Leads"),
        "redirect_leads": vals("Total Redirect Leads"), "redirect_verified": vals("Total Verified redirect leads"),
        "api_leads": vals("Total API Leads"), "api_verified": vals("Total Verified API leads"),
        "relevant_leads": vals("Relevant Leads"),
        "application_counts": {
            c: {"with_code": with_code.get(c, 0), "without_code": without_code.get(c, 0),
               "via_redirect": via_redirect.get(c, 0), "via_api": via_api.get(c, 0)}
            for c in columns
        },
        "amount_spent": vals("Amount Spent"), "additional_attributed": vals("Additional Attributed Applications"),
    }


def _zero_result(columns: List[str], stage_rows: List[str]) -> Dict[str, Any]:
    """A build_result()-shaped dict with every count at 0, for a column set a
    snapshot doesn't actually have data for (e.g. a program added to settings
    after the latest report was computed)."""
    zero = {c: 0 for c in columns}
    return build_result(
        programs=columns, stage_rows=stage_rows,
        matrix_counts={row: dict(zero) for row in stage_rows},
        total_leads=zero, verified_leads=zero, redirect_leads=zero, redirect_verified=zero,
        api_leads=zero, api_verified=zero, relevant_leads=zero,
        application_counts={c: {} for c in columns}, amount_spent=zero, additional_attributed=zero,
        detected_columns={}, data_quality={},
    )


def _diff_counts(end_vals: Dict[str, float], start_vals: Dict[str, float]) -> Dict[str, float]:
    """(end - start) per key, clamped at 0 - a count should never go negative
    between two cumulative snapshots taken further apart in time."""
    keys = set(end_vals) | set(start_vals)
    return {k: max(0, (end_vals.get(k, 0) or 0) - (start_vals.get(k, 0) or 0)) for k in keys}


def _diff_result(end_result: Dict[str, Any], start_result: Optional[Dict[str, Any]],
                 columns: List[str], stage_rows: List[str],
                 money: Optional[tuple] = None) -> Dict[str, Any]:
    """Rebuilds a build_result()-shaped dict as (end snapshot - start snapshot),
    reusing build_result() itself so every percentage/total stays internally
    consistent with how a single freshly-computed report looks. `money`, if
    given, overrides amount_spent/additional_attributed with pre-summed values -
    those are typed in per week in Generate Report, not part of the cumulative
    CRM export, so they're summed across the window elsewhere rather than diffed."""
    end_raw = _extract_raw_counts(end_result, columns, stage_rows)
    zero = {c: 0 for c in columns}
    start_raw = (_extract_raw_counts(start_result, columns, stage_rows) if start_result else
                {"matrix_counts": {row: zero for row in stage_rows},
                 "total_leads": zero, "verified_leads": zero, "redirect_leads": zero,
                 "redirect_verified": zero, "api_leads": zero, "api_verified": zero,
                 "relevant_leads": zero,
                 "application_counts": {c: {"with_code": 0, "without_code": 0, "via_redirect": 0, "via_api": 0} for c in columns},
                 "amount_spent": zero, "additional_attributed": zero})

    matrix_counts = {row: _diff_counts(end_raw["matrix_counts"][row], start_raw["matrix_counts"][row]) for row in stage_rows}
    application_counts = {
        c: {k: max(0, (end_raw["application_counts"][c][k] or 0) - (start_raw["application_counts"][c][k] or 0))
           for k in ("with_code", "without_code", "via_redirect", "via_api")}
        for c in columns
    }
    amount_spent, additional_attributed = money if money else (
        _diff_counts(end_raw["amount_spent"], start_raw["amount_spent"]),
        _diff_counts(end_raw["additional_attributed"], start_raw["additional_attributed"]),
    )

    return build_result(
        programs=columns, stage_rows=stage_rows, matrix_counts=matrix_counts,
        total_leads=_diff_counts(end_raw["total_leads"], start_raw["total_leads"]),
        verified_leads=_diff_counts(end_raw["verified_leads"], start_raw["verified_leads"]),
        redirect_leads=_diff_counts(end_raw["redirect_leads"], start_raw["redirect_leads"]),
        redirect_verified=_diff_counts(end_raw["redirect_verified"], start_raw["redirect_verified"]),
        api_leads=_diff_counts(end_raw["api_leads"], start_raw["api_leads"]),
        api_verified=_diff_counts(end_raw["api_verified"], start_raw["api_verified"]),
        relevant_leads=_diff_counts(end_raw["relevant_leads"], start_raw["relevant_leads"]),
        application_counts=application_counts, amount_spent=amount_spent, additional_attributed=additional_attributed,
        detected_columns=end_result.get("detected_columns", {}), data_quality=dict(end_result.get("data_quality", {})),
    )


def aggregate_reports(reports, settings, start: Optional[str] = None, end: Optional[str] = None,
                      no_baseline: bool = False):
    """Each stored report is a full cumulative-to-date CRM export as of its
    week_date, not a delta of that week's new leads/applications - confirmed
    directly with the user, and reproduced (Total Leads tripling across 3
    re-uploads of the same file). Summing raw counts across multiple reports
    therefore multiply-counts the same lead/application once per week it was
    re-uploaded. Instead: lead/application/verified counts come from the single
    latest report at or before `end` (or the overall latest if `end` is None),
    diffed against the latest report strictly before `start` so a date-range
    query shows only what changed in that window rather than the full running
    total. Amount Spent / Additional Attributed are typed in per week (not part
    of the cumulative export) and are summed across every report whose week
    falls in the window, same as before this fix.

    `no_baseline=True` skips the pre-start diff entirely and just shows the
    latest-in-range snapshot as-is. Use this for a saved Season: a season's
    `start` marks where a self-contained admissions cycle began, not a rolling
    window inside one - the closest report before that start could easily be a
    *different, unrelated* season's final snapshot (e.g. last year's), and
    diffing against it would silently subtract one season's totals from
    another's. Rolling presets ("Last 4 weeks") inside a single ongoing season
    still want the real diff, so this stays opt-in rather than the default."""
    cfg = {**DEFAULT_SETTINGS, **(settings or {})}
    raw_programs = cfg["programs"]
    merge_target = program_merge_targets(raw_programs)
    programs = [p for p in raw_programs if merge_target[p] == p]
    stage_rows = cfg["stage_rows"]

    ready = sorted((r for r in reports if r.get("result")), key=lambda r: r.get("week_date") or "")
    end_candidates = [r for r in ready if not end or (r.get("week_date") or "") <= end]
    start_candidates = [] if no_baseline else [r for r in ready if start and (r.get("week_date") or "") < start]
    snapshot_end = end_candidates[-1] if end_candidates else None
    snapshot_start = start_candidates[-1] if start_candidates else None
    window = [r for r in ready if (not start or (r.get("week_date") or "") >= start)
                              and (not end or (r.get("week_date") or "") <= end)]
    weeks = len(window)

    if snapshot_end is None:
        zero = {p: 0 for p in programs}
        result = build_result(
            programs=programs, stage_rows=stage_rows,
            matrix_counts={row: dict(zero) for row in stage_rows},
            total_leads=zero, verified_leads=zero, redirect_leads=zero, redirect_verified=zero,
            api_leads=zero, api_verified=zero, relevant_leads=zero,
            application_counts={p: {} for p in programs}, amount_spent=zero, additional_attributed=zero,
            detected_columns={}, data_quality={},
        )
        result["data_quality"] = {"weeks_aggregated": 0}
        return result

    amount_spent = {p: 0.0 for p in programs}
    additional_attributed = {p: 0.0 for p in programs}
    for r in window:
        for p, v in (r.get("amount_spent") or {}).items():
            if p in amount_spent:
                amount_spent[p] += float(v or 0)
        for p, v in (r.get("additional_attributed") or {}).items():
            if p in additional_attributed:
                additional_attributed[p] += float(v or 0)

    end_res, start_res = snapshot_end["result"], snapshot_start["result"] if snapshot_start else None
    result = _diff_result(end_res, start_res, programs, stage_rows, money=(amount_spent, additional_attributed))
    result["data_quality"]["weeks_aggregated"] = weeks

    # Publisher cumulative (overall + per program): diff the same two snapshots. A
    # program missing from a snapshot (e.g. settings changed since that report was
    # computed) falls back to zero-filled data on the "All" publisher columns,
    # rather than disappearing from the output entirely.
    def pub_diff(getter, fallback_cols: List[str]):
        e = getter(end_res)
        s = getter(start_res) if start_res else None
        cols = (e or {}).get("programs") or (s or {}).get("programs") or fallback_cols
        if not cols:
            return None
        return _diff_result(e or _zero_result(cols, stage_rows), s, cols, stage_rows)

    overall = pub_diff(lambda res: res.get("publisher_report"), [])
    if overall:
        result["publisher_report"] = overall
        pub_cols = overall.get("programs", [])
        pub_reports = {"All": overall}
        for prog_name in programs:
            pr = pub_diff(lambda res, p=prog_name: (res.get("publisher_reports") or {}).get(p), pub_cols)
            if pr:
                pub_reports[prog_name] = pr
        result["publisher_reports"] = pub_reports

    # Program-per-publisher: diff each publisher's sub-result across the same two snapshots.
    pub_keys = set((end_res.get("program_reports") or {}).keys())
    if start_res:
        pub_keys |= set((start_res.get("program_reports") or {}).keys())
    pub_keys.discard("All")
    prog_reports = {"All": {k: result[k] for k in
                    ("programs", "columns", "matrix", "summary") if k in result}}
    for pub_name in pub_keys:
        e = (end_res.get("program_reports") or {}).get(pub_name)
        if not e:
            continue
        s = (start_res.get("program_reports") or {}).get(pub_name) if start_res else None
        prog_reports[pub_name] = _diff_result(e, s, programs, stage_rows)
    if len(prog_reports) > 1:
        result["program_reports"] = prog_reports

    # Geography: diff matching locations between the two snapshots, per program x
    # publisher slice (both "All" for the plain per-program/per-publisher views,
    # and paired for the "one publisher within one program" view).
    def geo_diff_locations(e: Dict[str, Any], s: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
        merged: Dict[str, Dict[str, Any]] = {}
        for loc in set(e) | set(s):
            ev, sv = e.get(loc, {}), s.get(loc, {})
            leads = max(0, (ev.get("leads", 0) or 0) - (sv.get("leads", 0) or 0))
            verified = max(0, (ev.get("verified_leads", 0) or 0) - (sv.get("verified_leads", 0) or 0))
            relevant = max(0, (ev.get("relevant_leads", 0) or 0) - (sv.get("relevant_leads", 0) or 0))
            apps = max(0, (ev.get("applications", 0) or 0) - (sv.get("applications", 0) or 0))
            if leads or verified or relevant or apps:
                merged[loc] = {"leads": leads, "verified_leads": verified, "relevant_leads": relevant,
                              "applications": apps,
                              "conversion_pct": round(apps / leads * 100, 2) if leads else None}
        return merged

    def geo_diff(key: str) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        end_geo = end_res.get(key) or {}
        start_geo = (start_res.get(key) or {}) if start_res else {}
        for prog_key in set(end_geo) | set(start_geo):
            e_pubs = end_geo.get(prog_key) or {}
            s_pubs = start_geo.get(prog_key) or {}
            pub_out: Dict[str, Any] = {}
            for pub_key in set(e_pubs) | set(s_pubs):
                merged = geo_diff_locations(e_pubs.get(pub_key) or {}, s_pubs.get(pub_key) or {})
                if merged:
                    pub_out[pub_key] = merged
            if pub_out:
                out[prog_key] = pub_out
        return out

    geo_state = geo_diff("geo_by_state")
    if geo_state:
        result["geo_by_state"] = geo_state
    geo_city = geo_diff("geo_by_city")
    if geo_city:
        result["geo_by_city"] = geo_city
    return result
