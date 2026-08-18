"""E2E coverage for coupon counts, dynamic publisher/course settings, dimensions, and report regressions."""
import io
import os
import time
from pathlib import Path

import pytest
import requests
from dotenv import dotenv_values
from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
frontend_env = dotenv_values(REPO_ROOT / "frontend" / ".env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL is missing")
BASE_URL = base_url.rstrip("/")

# Test-run credentials for a user that must already exist in the target DB
# (e.g. the bootstrapped admin — see backend/.env.example ADMIN_USERNAME/ADMIN_PASSWORD).
TEST_USERNAME = os.environ.get("TEST_ADMIN_USERNAME") or os.environ.get("ADMIN_USERNAME")
TEST_PASSWORD = os.environ.get("TEST_ADMIN_PASSWORD") or os.environ.get("ADMIN_PASSWORD")
if not TEST_USERNAME or not TEST_PASSWORD:
    raise RuntimeError(
        "TEST_ADMIN_USERNAME/TEST_ADMIN_PASSWORD (or ADMIN_USERNAME/ADMIN_PASSWORD) "
        "must be set to a valid login for the target backend."
    )

DEFAULT_PROGRAMS = ["B.Com", "BBA", "PGDM"]
EXPECTED_PUBLISHERS = {
    "Collegedunia", "Shiksha", "Google Ads", "Meta Ads", "CollegeSearch",
    "GetMyUni", "CareerS360", "NVT", "Organic", "Referral", "Telephony Inbound",
}
STATE = {"created_ids": []}


def workbook_bytes(headers, rows):
    """Create a small in-memory XLSX fixture."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def summary_row(result, label):
    """Return a named report summary row."""
    return next(row for row in result["summary"] if row.get("label") == label)


def wait_until_ready(client, report_id, timeout=60):
    """Poll asynchronous report generation to a terminal ready state."""
    deadline = time.time() + timeout
    latest = None
    while time.time() < deadline:
        response = client.get(f"{BASE_URL}/api/reports/{report_id}", timeout=20)
        assert response.status_code == 200, response.text
        latest = response.json()
        if latest.get("status") == "ready":
            return latest
        if latest.get("status") == "error":
            pytest.fail(f"Report processing failed: {latest.get('error')}")
        time.sleep(0.5)
    pytest.fail(f"Report {report_id} did not become ready; latest={latest}")


def create_sample(client):
    """Create, track, and wait for one fresh sample report."""
    response = client.post(f"{BASE_URL}/api/reports/sample", timeout=20)
    assert response.status_code == 200, response.text
    created = response.json()
    assert created["status"] == "processing"
    assert isinstance(created["id"], str) and created["id"]
    STATE["created_ids"].append(created["id"])
    return wait_until_ready(client, created["id"])


def upload_report(client, label, lead_bytes, application_files=None):
    """Upload a report with optional multiple application workbooks."""
    files = [("lead_file", ("TEST_leads.xlsx", lead_bytes,
             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))]
    for index, content in enumerate(application_files or []):
        files.append(("application_files", (f"TEST_apps_{index}.xlsx", content,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")))
    response = client.post(
        f"{BASE_URL}/api/reports",
        data={"week_label": label, "week_date": "2026-07-15", "amount_spent": "{}",
              "additional_attributed": "{}"},
        files=files,
        timeout=30,
    )
    assert response.status_code == 200, response.text
    created = response.json()
    STATE["created_ids"].append(created["id"])
    return wait_until_ready(client, created["id"])


@pytest.fixture(scope="module")
def api_client():
    """Authenticated API client with deterministic settings and cleanup."""
    session = requests.Session()
    session.headers.update({"Accept": "application/json"})
    login = session.post(
        f"{BASE_URL}/api/auth/login",
        json={"username": TEST_USERNAME, "password": TEST_PASSWORD},
        timeout=20,
    )
    assert login.status_code == 200, login.text
    token = login.json()["access_token"]
    session.headers.update({"Authorization": f"Bearer {token}"})
    reset = session.put(
        f"{BASE_URL}/api/settings",
        json={"programs": DEFAULT_PROGRAMS, "included_publishers": [],
              "excluded_publishers": [], "exclude_test_leads": True,
              "test_keywords": ["test"]},
        timeout=20,
    )
    assert reset.status_code == 200, reset.text
    yield session
    # Required final state: defaults and all detected publishers.
    session.put(
        f"{BASE_URL}/api/settings",
        json={"programs": DEFAULT_PROGRAMS, "included_publishers": [],
              "excluded_publishers": [], "exclude_test_leads": True,
              "test_keywords": ["test"]},
        timeout=20,
    )
    for report_id in STATE["created_ids"]:
        session.delete(f"{BASE_URL}/api/reports/{report_id}", timeout=20)
    session.close()


class TestLeadPulseDynamicDimensions:
    """Coupon parsing plus complete and selectable publisher/course dimensions."""

    def test_01_discount_coupon_counts_blank_tokens_as_without_code(self, api_client):
        lead_bytes = workbook_bytes(
            ["Registered Name", "Course", "Lead Stage", "Publisher Name",
             "Email Verification Status", "Mobile Verification Status"],
            [["Real One", "PGDM", "APPLIED", "NVT", "VERIFIED", ""],
             ["Real Two", "BBA", "WARM", "Organic", "", "VERIFIED"],
             ["Real Three", "B.Com", "COLD", "Referral", "", ""]],
        )
        discount_coupon = workbook_bytes(
            ["Programme", "Discount Coupon"],
            [["PGDM", "SAVE10"], ["PGDM", None], ["PGDM", "NA"], ["PGDM", 0],
             ["BBA", "BBA10"], ["BBA", "N/A"]],
        )
        coupon_code = workbook_bytes(
            ["Programme", "Coupon Code"],
            [["BBA", "BBA20"], ["BBA", "-"], ["B.Com", " PROMO "], ["B.Com", "nil"]],
        )
        report = upload_report(
            api_client, "TEST_Discount coupon parsing", lead_bytes,
            [discount_coupon, coupon_code],
        )
        result = report["result"]
        with_codes = summary_row(result, "No. of Applications with codes")
        without_codes = summary_row(result, "No. of Applications without codes")
        assert with_codes["values"] == {"B.Com": 1, "BBA": 2, "PGDM": 1}
        assert with_codes["total"] == 4
        assert without_codes["values"] == {"B.Com": 1, "BBA": 2, "PGDM": 3}
        assert without_codes["total"] == 6
        persisted = api_client.get(f"{BASE_URL}/api/reports/{report['id']}", timeout=20)
        assert persisted.status_code == 200
        assert persisted.json()["application_counts"]["PGDM"] == {
            "with_code": 1, "without_code": 3, "via_redirect": 0, "via_api": 0,
        }

    def test_02_sample_has_all_publishers_and_available_dimensions(self, api_client):
        report = create_sample(api_client)
        STATE["default_sample_id"] = report["id"]
        result = report["result"]
        publishers = result["publisher_report"]["programs"]
        assert set(publishers) == EXPECTED_PUBLISHERS
        assert len(publishers) == 11
        assert "NVT" in publishers
        assert "Sulekha" not in publishers
        assert "Other" not in publishers
        assert result["publisher_report"]["columns"] == publishers + ["Total"]

        available_publishers = result["data_quality"]["available_publishers"]
        available_courses = result["data_quality"]["available_courses"]
        assert {item["name"] for item in available_publishers} == EXPECTED_PUBLISHERS
        assert all(isinstance(item["count"], int) and item["count"] > 0 for item in available_publishers)
        assert {item["name"] for item in available_courses} >= set(DEFAULT_PROGRAMS)
        assert all(isinstance(item["count"], int) and item["count"] > 0 for item in available_courses)

        applied = next(row for row in result["matrix"] if row["stage"] == "APPLIED")
        assert applied["values"] == {"B.Com": 14, "BBA": 41, "PGDM": 71}

    def test_03_included_and_excluded_publishers_apply_to_new_reports(self, api_client):
        selected = api_client.put(
            f"{BASE_URL}/api/settings",
            json={"included_publishers": ["Collegedunia", "NVT"], "excluded_publishers": []},
            timeout=20,
        )
        assert selected.status_code == 200
        assert selected.json()["included_publishers"] == ["Collegedunia", "NVT"]
        selected_report = create_sample(api_client)
        assert selected_report["result"]["publisher_report"]["programs"] == ["Collegedunia", "NVT"]
        assert set(selected_report["result"]["program_reports"]) == {"All", "Collegedunia", "NVT"}

        excluded = api_client.put(
            f"{BASE_URL}/api/settings",
            json={"included_publishers": [], "excluded_publishers": ["NVT"]},
            timeout=20,
        )
        assert excluded.status_code == 200
        excluded_report = create_sample(api_client)
        excluded_names = set(excluded_report["result"]["publisher_report"]["programs"])
        assert excluded_names == EXPECTED_PUBLISHERS - {"NVT"}

        reset = api_client.put(
            f"{BASE_URL}/api/settings",
            json={"included_publishers": [], "excluded_publishers": []},
            timeout=20,
        )
        assert reset.status_code == 200
        reset_report = create_sample(api_client)
        assert set(reset_report["result"]["publisher_report"]["programs"]) == EXPECTED_PUBLISHERS

    def test_04_selected_courses_apply_with_generic_matching_and_totals(self, api_client):
        changed = api_client.put(
            f"{BASE_URL}/api/settings", json={"programs": ["PGDM", "BBA"]}, timeout=20
        )
        assert changed.status_code == 200
        assert changed.json()["programs"] == ["PGDM", "BBA"]
        report = create_sample(api_client)
        result = report["result"]
        assert result["programs"] == ["PGDM", "BBA"]
        assert result["columns"] == ["PGDM", "BBA", "Total"]
        applied = next(row for row in result["matrix"] if row["stage"] == "APPLIED")
        assert applied["values"] == {"PGDM": 71, "BBA": 41}
        assert applied["total"] == 112
        total_leads = summary_row(result, "Total Leads")
        assert total_leads["total"] == sum(total_leads["values"].values())

        reset = api_client.put(
            f"{BASE_URL}/api/settings", json={"programs": DEFAULT_PROGRAMS}, timeout=20
        )
        assert reset.status_code == 200
        assert reset.json()["programs"] == DEFAULT_PROGRAMS

    def test_05_available_endpoint_aggregates_ready_report_dimensions(self, api_client):
        response = api_client.get(f"{BASE_URL}/api/available", timeout=30)
        assert response.status_code == 200
        data = response.json()
        assert set(data) == {"courses", "publishers"}
        assert set(DEFAULT_PROGRAMS).issubset({item["name"] for item in data["courses"]})
        assert EXPECTED_PUBLISHERS.issubset({item["name"] for item in data["publishers"]})
        assert all(set(item) == {"name", "count"} for item in data["courses"] + data["publishers"])
        assert all(isinstance(item["count"], int) and item["count"] > 0
                   for item in data["courses"] + data["publishers"])
    def test_06_report_variants_and_amount_patches_persist(self, api_client):
        report_id = STATE["default_sample_id"]
        response = api_client.get(f"{BASE_URL}/api/reports/{report_id}", timeout=20)
        assert response.status_code == 200
        report = response.json()
        result = report["result"]
        assert set(result["publisher_reports"]) == {"All", *DEFAULT_PROGRAMS}
        assert set(result["program_reports"]) == {"All", *EXPECTED_PUBLISHERS}
        for nested in result["publisher_reports"].values():
            assert nested["columns"] == nested["programs"] + ["Total"]
            assert len(nested["matrix"]) == 10
        for nested in result["program_reports"].values():
            assert nested["programs"] == DEFAULT_PROGRAMS
            assert nested["columns"] == DEFAULT_PROGRAMS + ["Total"]

        amounts = {"B.Com": 1000, "BBA": 2000, "PGDM": 3000}
        attributed = {"B.Com": 1, "BBA": 2, "PGDM": 3}
        patched = api_client.patch(
            f"{BASE_URL}/api/reports/{report_id}/amounts",
            json={"amount_spent": amounts, "additional_attributed": attributed}, timeout=30,
        )
        assert patched.status_code == 200, patched.text
        patched_doc = patched.json()
        assert summary_row(patched_doc["result"], "Amount Spent")["values"] == amounts
        assert patched_doc["kpis"]["amount_spent"] == 6000

        publisher = "Collegedunia"
        pub_patched = api_client.patch(
            f"{BASE_URL}/api/reports/{report_id}/publisher-amounts",
            json={"amount_spent": {publisher: 4321}, "cpa": {}}, timeout=30,
        )
        assert pub_patched.status_code == 200, pub_patched.text
        pub_doc = pub_patched.json()
        assert summary_row(pub_doc["result"]["publisher_report"], "Amount Spent")["values"][publisher] == 4321
        assert pub_doc["result"]["publisher_reports"]["All"] == pub_doc["result"]["publisher_report"]
        persisted = api_client.get(f"{BASE_URL}/api/reports/{report_id}", timeout=20).json()
        assert persisted["result"]["publisher_reports"]["All"] == persisted["result"]["publisher_report"]

    def test_07_cumulative_ranges_export_and_nested_reports(self, api_client):
        response = api_client.get(f"{BASE_URL}/api/reports/cumulative", timeout=90)
        assert response.status_code == 200, response.text
        doc = response.json()
        assert doc["status"] == "ready"
        assert doc["result"]["data_quality"]["weeks_aggregated"] >= 2
        assert doc["kpis"]["total_leads"] > 0
        assert set(doc["result"]["publisher_reports"]) == {"All", *DEFAULT_PROGRAMS}
        assert "All" in doc["result"]["program_reports"]

        malformed = api_client.get(
            f"{BASE_URL}/api/reports/cumulative", params={"start": "bad-date"}, timeout=30
        )
        assert malformed.status_code == 400
        assert "Invalid start date" in malformed.json()["detail"]
        reversed_dates = api_client.get(
            f"{BASE_URL}/api/reports/cumulative",
            params={"start": "2026-12-31", "end": "2026-01-01"}, timeout=30,
        )
        assert reversed_dates.status_code == 400
        assert "start date must be on or before end date" in reversed_dates.json()["detail"]

        exported = api_client.get(
            f"{BASE_URL}/api/reports/cumulative/export", params={"start": "2026-01-01"}, timeout=90
        )
        assert exported.status_code == 200
        assert exported.content[:2] == b"PK"
        workbook = load_workbook(io.BytesIO(exported.content), read_only=True, data_only=True)
        # One "Weekly Report" summary sheet plus one "Publisher - <key>" sheet per
        # publisher_reports entry (All + each program) — see excel_export.build_workbook.
        assert set(workbook.sheetnames) == {"Weekly Report", "Publisher - All",
                                            *(f"Publisher - {p}" for p in DEFAULT_PROGRAMS)}

    def test_08_single_export_trends_and_missing_report_errors(self, api_client):
        report_id = STATE["default_sample_id"]
        exported = api_client.get(f"{BASE_URL}/api/reports/{report_id}/export", timeout=90)
        assert exported.status_code == 200
        assert exported.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert exported.content[:2] == b"PK"

        trends = api_client.get(f"{BASE_URL}/api/trends", timeout=30)
        assert trends.status_code == 200
        trend_data = trends.json()
        assert isinstance(trend_data, list) and trend_data
        assert [item["week_date"] for item in trend_data] == sorted(item["week_date"] for item in trend_data)
        assert report_id in {item["id"] for item in trend_data}
        assert all(isinstance(item.get("kpis"), dict) for item in trend_data)

        missing = api_client.get(f"{BASE_URL}/api/reports/not-a-real-report", timeout=20)
        assert missing.status_code == 404
        assert missing.json()["detail"] == "Report not found"

    def test_09_test_leads_remain_excluded(self, api_client):
        lead_bytes = workbook_bytes(
            ["Registered Name", "Course", "Lead Stage", "Email", "Lead Remark",
             "Publisher Name", "Email Verification Status", "Mobile Verification Status"],
            [["Test User", "PGDM", "APPLIED", "real@example.com", "", "NVT", "VERIFIED", ""],
             ["Real User", "BBA", "WARM", "real2@example.com", "", "Organic", "", "VERIFIED"],
             ["Another", "B.Com", "TEST LEADS", "real3@example.com", "", "Referral", "", ""],
             ["Valid User", "PGDM", "APPLIED", "valid@example.com", "", "NVT", "VERIFIED", ""]],
        )
        report = upload_report(api_client, "TEST_Test lead exclusion", lead_bytes)
        quality = report["result"]["data_quality"]
        assert quality["raw_rows"] == 4
        assert quality["test_leads_excluded"] == 2
        assert quality["total_rows"] == 2
        assert report["kpis"]["total_leads"] == 2

    def test_10_final_settings_are_default(self, api_client):
        reset = api_client.put(
            f"{BASE_URL}/api/settings",
            json={"programs": DEFAULT_PROGRAMS, "included_publishers": [],
                  "excluded_publishers": []}, timeout=20,
        )
        assert reset.status_code == 200
        data = reset.json()
        assert data["programs"] == DEFAULT_PROGRAMS
        assert data["included_publishers"] == []
        assert data["excluded_publishers"] == []
