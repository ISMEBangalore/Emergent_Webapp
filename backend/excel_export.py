"""Build a styled .xlsx report from a computed report result."""
import io
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GREEN = "C6EFCE"
YELLOW = "FFF2CC"
BLUE_HDR = "9DC3E6"
BLUE_ROW = "BDD7EE"
CYAN = "00B0F0"
SECTION = "FFE699"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill(hexcolor):
    return PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")


def _sheet_name(name: str) -> str:
    """Excel sheet names must be <=31 chars and can't contain []:*?/\\."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    return name[:31] or "Sheet"


def _is_empty_breakdown(result: Dict[str, Any]) -> bool:
    programs = (result or {}).get("programs") or []
    return not programs or (len(programs) == 1 and programs[0] == "Unknown")


def _total_applications(result: Dict[str, Any]) -> float:
    for s in (result or {}).get("summary") or []:
        if s.get("label") == "Total No. of Applications":
            return s.get("total") or 0
    return 0


def _write_report_sheet(ws, result: Dict[str, Any], title: str) -> None:
    programs = result["programs"]
    ncols = 2 + len(programs) * 2 + 1  # label + (prog,%)*n + Total

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    # Header row
    r = 2
    headers = ["Lead Stage"]
    for p in programs:
        headers += [p, "%"]
    headers += ["Total"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = _fill(BLUE_HDR if i == 1 else GREEN)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def pctstr(v):
        return f"{round(v)}%" if v is not None else "-"

    # Matrix rows
    for m in result["matrix"]:
        r += 1
        ws.cell(row=r, column=1, value=m["stage"]).border = border
        col = 2
        for p in programs:
            vc = ws.cell(row=r, column=col, value=m["values"][p]); vc.fill = _fill(YELLOW); vc.border = border
            pc = ws.cell(row=r, column=col + 1, value=pctstr(m["pct"][p])); pc.fill = _fill(GREEN); pc.border = border
            pc.alignment = Alignment(horizontal="center")
            col += 2
        ws.cell(row=r, column=col, value=m["total"]).border = border

    # Summary rows
    for s in result["summary"]:
        r += 1
        if s.get("kind") == "section":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            cell = ws.cell(row=r, column=1, value=s["label"])
            cell.font = Font(bold=True, italic=True, color="C00000")
            cell.alignment = Alignment(horizontal="center")
            continue
        fmt = s.get("fmt", "int")
        label_cell = ws.cell(row=r, column=1, value=s["label"])
        label_cell.font = Font(bold=True)
        label_cell.fill = _fill(BLUE_ROW if s.get("kind") == "header" else YELLOW)
        label_cell.border = border
        col = 2
        for p in programs:
            if fmt == "pct_only":
                v = pctstr(s["pct"].get(p))
                cell = ws.cell(row=r, column=col, value=v)
                cell.alignment = Alignment(horizontal="center")
            elif fmt == "money":
                cell = ws.cell(row=r, column=col, value=s["values"][p])
                cell.number_format = '#,##0'
            else:
                cell = ws.cell(row=r, column=col, value=s["values"][p])
            cell.border = border
            # percentage column
            pc = ws.cell(row=r, column=col + 1,
                         value=pctstr(s["pct"].get(p)) if (s.get("pct") and fmt != "pct_only") else "")
            pc.border = border
            pc.alignment = Alignment(horizontal="center")
            col += 2
        # total
        if fmt == "pct_only":
            tv = pctstr(s.get("total_pct"))
        else:
            tv = s.get("total")
        tcell = ws.cell(row=r, column=col, value=tv)
        tcell.border = border
        if fmt == "money":
            tcell.number_format = '#,##0'

    # column widths
    ws.column_dimensions["A"].width = 34
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.freeze_panes = "B3"


def build_workbook(doc: Dict[str, Any], top_publishers: int = 0) -> bytes:
    result = doc["result"]
    title = doc.get("week_label") or "Weekly CRM Report"

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Report"
    _write_report_sheet(ws, result, title)

    # One sheet per program's publisher breakdown (the "By Publisher" dashboard tab),
    # so the export covers what's shown on screen, not just the program summary.
    for key, pub in (result.get("publisher_reports") or {}).items():
        if _is_empty_breakdown(pub):
            continue
        ws2 = wb.create_sheet(_sheet_name(f"Publisher - {key}"))
        _write_report_sheet(ws2, pub, f"By Publisher — {key}")

    # Optional: program breakdown for the top N publishers by application volume
    # (the "Programs per Publisher" dashboard tab). Unbounded by default, opt-in.
    if top_publishers > 0:
        prog_reports = result.get("program_reports") or {}
        candidates = [(k, v) for k, v in prog_reports.items() if k != "All" and not _is_empty_breakdown(v)]
        candidates.sort(key=lambda kv: -_total_applications(kv[1]))
        for pub_name, prog in candidates[:top_publishers]:
            ws3 = wb.create_sheet(_sheet_name(f"Programs - {pub_name}"))
            _write_report_sheet(ws3, prog, f"Programs per Publisher — {pub_name}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
